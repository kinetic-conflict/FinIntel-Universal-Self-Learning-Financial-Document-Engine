from openpyxl import load_workbook
from datetime import datetime, date
from decimal import Decimal

from app.jobs import update_job_status, set_job_result, get_job
from app.client import call_ocr_service, call_validation_service


def guess_doc_type(filename: str) -> str:
    name = filename.lower()

    if "bank" in name or "statement" in name:
        return "bank_statement"
    if "payslip" in name or "salary" in name:
        return "payslip"
    if "invoice" in name or "bill" in name:
        return "invoice"

    return "unknown"


def is_excel(filename: str) -> bool:
    return filename.lower().endswith(".xlsx")


def make_json_safe(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def excel_to_knowledge(excel_path: str) -> dict:
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))

    if not rows or len(rows) < 2:
        return {"rows": []}

    headers = [str(h).strip() if h else "" for h in rows[0]]

    data_rows = []
    for row in rows[1:]:
        record = {}
        for i, header in enumerate(headers):
            key = header if header else f"col_{i}"
            value = row[i] if i < len(row) else None
            record[key] = make_json_safe(value)
        data_rows.append(record)

    return {"rows": data_rows}


def process_document(job_id: str):
    update_job_status(job_id, "RUNNING")

    try:
        job = get_job(job_id)
        if not job:
            set_job_result(job_id, {"error": "Job not found"})
            update_job_status(job_id, "FAILED")
            return

        filename = job["filename"]
        path = job["path"]

        doc_type = guess_doc_type(filename)

        if is_excel(filename):
            extracted = excel_to_knowledge(path)

            knowledge_object = {
                "source": "excel",
                "doc_type": doc_type,
                "entities": extracted,
                "tables": [],
                "metadata": {
                    "filename": filename,
                    "job_id": job_id
                }
            }
        else:
            ocr_result = call_ocr_service(path)

            knowledge_object = {
                "source": "ocr",
                "doc_type": doc_type,
                "entities": ocr_result.get("entities", {}),
                "tables": ocr_result.get("tables", []),
                "metadata": {
                    "filename": filename,
                    "job_id": job_id
                }
            }

        validation_result = call_validation_service(knowledge_object)

        result = {
            "doc_type": doc_type,
            "source": knowledge_object["source"],
            "entities": knowledge_object["entities"],
            "tables": knowledge_object["tables"],
            "validation": validation_result,
            "metadata": knowledge_object["metadata"]
        }

        set_job_result(job_id, result)
        update_job_status(job_id, "DONE")

    except Exception as e:
        set_job_result(job_id, {"error": str(e)})
        update_job_status(job_id, "FAILED")
